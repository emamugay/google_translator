#!/usr/bin/python
# -*- coding: utf-8 -*-

import os, sys, time
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from openpyxl import load_workbook

def start_new_session():
    options = Options()
    options.headless = False

    # Change the path to your geckodriver path
    DRIVER_PATH = "/mnt/Devs/Python/scrapy/mmi/driver/geckodriver"
    options.binary_location = "/usr/bin/firefox"

    driver = webdriver.Firefox(options=options, executable_path=DRIVER_PATH)
    driver.set_window_size(1920,1080)

    return driver

def end_current_session(driver):
    driver.quit()

def process_file(file_src = ""):
    if file_src != "":
        files = file_src
    else:
        print('No file selected')
        print('to translate, run python translate_prod.py <myfile.xlsx>')
        exit(1)
        
    wb = load_workbook(files)
    sheet = wb.active
    new_content = "Translated Contents"
    new_title = "Translated Title"

    sheet.insert_cols(5,1)
    sheet['E1'] = new_content

    sheet.insert_cols(20,1)
    sheet['T1'] = new_title

    #columns = sheet.max_column # Number of fields
    cells = sheet.max_row # Number of records

    driver = start_new_session()
    driver.get("https://www.google.com/search?q=google+translate")
    ss = driver.find_element_by_css_selector('#tw-source-text-ta')
    ss.clear()	

    words = []
    for i in range(2, cells):
        translated_content = ''
        words.clear()
        # Contents
        try:
            word_len = SplitNumber(sheet['D'+str(i)].value)
            if word_len > 300:
                words = SplitWords(sheet['D'+str(i)].value)
                for word in words:
                    ss = driver.find_element_by_css_selector('#tw-source-text-ta').send_keys(word)
                    time.sleep(2)
                    translated_content += ' ' + str(driver.find_element_by_css_selector('#tw-target-text').text)
                    driver.find_element_by_css_selector('#tw-source-text-ta').clear()	

                sheet['E'+str(i)] = translated_content
                ss = driver.find_element_by_css_selector('#tw-source-text-ta').send_keys(sheet['S'+str(i)].value)
                time.sleep(2)
                translated_content = str(driver.find_element_by_css_selector('#tw-target-text').text)
                sheet['T'+str(i)] = translated_content
                driver.find_element_by_css_selector('#tw-source-text-ta').clear()	
                    
            else:                
                ss = driver.find_element_by_css_selector('#tw-source-text-ta').send_keys(sheet['D'+str(i)].value)
                time.sleep(2)
                translated_content = str(driver.find_element_by_css_selector('#tw-target-text').text)
                sheet['E'+str(i)] = translated_content
                driver.find_element_by_css_selector('#tw-source-text-ta').clear()	

                ss = driver.find_element_by_css_selector('#tw-source-text-ta').send_keys(sheet['S'+str(i)].value)
                time.sleep(2)
                translated_content = str(driver.find_element_by_css_selector('#tw-target-text').text)
                sheet['T'+str(i)] = translated_content
                driver.find_element_by_css_selector('#tw-source-text-ta').clear()	

            print(f'Row : {i}')
            
        except BaseException as e:
            print("ERROR: "+str(e))
            continue

    file_name = os.path.basename(os.path.splitext(file_src)[0])
    file_ext = os.path.basename(os.path.splitext(file_src)[1])
    wb.save(file_name + '_Translated_File' + file_ext)

def SplitNumber(src):
    try:
        
        txt_len = src.split(' ')
        wlen = 0
        for word in txt_len:
            wlen += len(word)

        return wlen
    
    except BaseException as e:
        print("ERROR: "+str(e))
        
# split content source into 300 words length 
def SplitWords(src):
    try:
        
        txt_len = src.split(' ')
        wlen = 0
        words = ''
        wordsArray = []
        for word in txt_len:
            wlen += len(word)
            words += ''.join(word)
            if wlen > 300:
                wordsArray.append(words)
                words = ""
                wlen = 0
        return wordsArray
    
    except BaseException as e:
        print("ERROR: "+str(e))
    

if __name__ == "__main__":   
    if len(sys.argv) >= 2:
        for file in sys.argv[1:]:
            process_file(file)
    else:
        print('No file selected')
        print('to translate, run python translate_prod.py <myfile.xlsx>')       
